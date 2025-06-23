import streamlit as st
import pandas as pd
from io import BytesIO
import base64
import openpyxl
import re

st.set_page_config(layout="wide")
st.title("Excel to Python Code Generator (Extended Formula Support)")

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx", "xlsm"])
if uploaded_file:
    try:
        in_memory_file = BytesIO(uploaded_file.read())
        wb = openpyxl.load_workbook(in_memory_file, data_only=False)
        sheet_names = wb.sheetnames
        selected_sheet = st.selectbox("Choose a sheet", sheet_names)
        ws = wb[selected_sheet]

        # Load data-only version for preview
        data_only_wb = openpyxl.load_workbook(BytesIO(uploaded_file.getvalue()), data_only=True)
        df = pd.DataFrame(data_only_wb[selected_sheet].values)
        st.subheader("Preview of Sheet Data")
        st.dataframe(df)

        converted_lines = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula = cell.value
                    coordinate = cell.coordinate

                    # INDEX
                    index_match = re.match(r"=INDEX\(([^,]+),\s*(\d+)\)", formula)
                    if index_match:
                        cell_range, index = index_match.groups()
                        col_letter = re.findall(r"[A-Z]+", cell_range)[0]
                        row_nums = list(map(int, re.findall(r"\d+", cell_range)))
                        row_start = row_nums[0] - 1
                        row_end = row_nums[-1]
                        col_index = ord(col_letter.upper()) - 65
                        python_expr = f"value_{coordinate} = df.iloc[{row_start}:{row_end}, {col_index}].iloc[{int(index)-1}]"
                        converted_lines.append(python_expr)
                        continue

                    # OFFSET
                    offset_match = re.match(r"=OFFSET\(([A-Z]+)(\d+),\s*(\d+),\s*(\d+)\)", formula)
                    if offset_match:
                        base_col, base_row, row_offset, col_offset = offset_match.groups()
                        base_row = int(base_row) - 1 + int(row_offset)
                        base_col_index = ord(base_col.upper()) - 65 + int(col_offset)
                        python_expr = f"value_{coordinate} = df.iloc[{base_row}, {base_col_index}]"
                        converted_lines.append(python_expr)
                        continue

                    # XLOOKUP
                    xlookup_match = re.match(r'=XLOOKUP\(([^,]+),\s*([A-Z]+\d+:[A-Z]+\d+),\s*([A-Z]+\d+:[A-Z]+\d+),\s*"([^"]+)"\)', formula)
                    if xlookup_match:
                        lookup_val, lookup_range, return_range, fallback = xlookup_match.groups()
                        lookup_col = re.findall(r"[A-Z]+", lookup_range)[0]
                        return_col = re.findall(r"[A-Z]+", return_range)[0]
                        lookup_col_index = ord(lookup_col.upper()) - 65
                        return_col_index = ord(return_col.upper()) - 65
                        python_expr = (
                            f"value_{coordinate} = df[df.iloc[:, {lookup_col_index}] == {lookup_val}].iloc[0, {return_col_index}] "
                            f"if {lookup_val} in df.iloc[:, {lookup_col_index}].values else '{fallback}'"
                        )
                        converted_lines.append(python_expr)
                        continue

                    converted_lines.append(f"# Could not parse formula in {coordinate}: {formula}")

        final_code = "import pandas as pd\n"
        final_code += f"df = pd.read_excel('your_file.xlsx', sheet_name='{selected_sheet}')\n"
        final_code += "\n".join(converted_lines) + "\n"
        final_code += "print(df.head())\n"

        st.subheader("Generated Python Code")
        st.code(final_code, language="python")

        b64 = base64.b64encode(final_code.encode()).decode()
        href = f'<a href="data:file/txt;base64,{b64}" download="generated_code.py">📥 Download Python Code</a>'
        st.markdown(href, unsafe_allow_html=True)

    except Exception as e:
        st.error(f"❌ Failed to process the Excel file: {e}")
