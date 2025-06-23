import pandas as pd
from openpyxl import load_workbook

def read_excel_and_generate_code(file_path):
    wb = load_workbook(filename=file_path, data_only=False)
    sheet = wb.active

    df = pd.read_excel(file_path)  # Reads actual values
    columns = df.columns.tolist()

    formulas = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[cell.coordinate] = cell.value

    print("📌 Columns:", columns)
    print("📌 Formulas Found:")
    for coord, formula in formulas.items():
        print(f"{coord}: {formula}")
    
    # Generate placeholder Python code
    python_code = "def calculate(data):\n"
    for col in columns:
        python_code += f"    data['{col}_processed'] = data['{col}']  # Placeholder logic\n"
    python_code += "    return data\n"

    with open("generated_logic.py", "w") as f:
        f.write(python_code)
    
    print("\n✅ Python code generated in 'generated_logic.py'.")

# Run the function on your Excel file
read_excel_and_generate_code("your_excel_file.xlsx")  # Replace this with your file name
